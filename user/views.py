from django.shortcuts import render, redirect
from .forms import UserLoginForm, UserRegisterForm, UserUpdateForm, ProfileUpdateForm,CircleUpdateForm
from django.contrib.auth import authenticate,login,logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .models import Circle
from Socialbook import settings
from django.core.mail import send_mail
from django.contrib.auth.models import User
from social.models import Post
from openpyxl import Workbook
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger





# Authentication System

def register(request):
    if request.method == 'POST':
       form = UserRegisterForm(request.POST)
       if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            email = form.cleaned_data.get('email')
            messages.success(request, f'Account created for {username}, You can now login.')


            return redirect('login')
    else:
        form = UserRegisterForm()
    return render(request,'user/register.html', {'form':form})

def login_view(request):
    if request.method == "POST":
        form = UserLoginForm(request.POST)
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username , password=password)
        if user is not None:
            login(request, user)
            return redirect("social-home")
        else:
            print(messages.warning(request,'Sorry, your password/username was incorrect. Please try again.'))
            return redirect("login")
    else:
        form = UserLoginForm()
    return render(request, "user/login.html", {"form": form})
        
@login_required        
def logout_view(request):
    messages.info(request,f"{request.user.username} has been successfully logged out.")
    logout(request)
    return render(request, "user/logout.html")


    

# Creating & Updating Profile 

@login_required 
def profile(request):
    if request.method == "POST": 
       u_form = UserUpdateForm(request.POST,instance=request.user)
       p_form = ProfileUpdateForm(request.POST,request.FILES,instance=request.user.profile)
       if u_form.is_valid() and p_form.is_valid():
           u_form.save()
           p_form.save()
           print(messages.success(request,'Your account has been updated.'))
           return redirect("profile")

    else:
        u_form = UserUpdateForm(instance=request.user)
        p_form = ProfileUpdateForm(instance=request.user.profile)     

    context = {
        'u_form' : u_form ,
        'p_form' : p_form
    }

    return render(request, 'user/profile.html',context)




# Personalised feed

@login_required 
def feed(request):

    posts_list = []
    for friend in Circle.objects.get(owner=request.user).friends.all():
          for post in Post.objects.filter(author=friend):
            posts_list.append(post)

    page = request.GET.get('page', 1)
    paginator = Paginator(posts_list, 5)

    try:
        posts = paginator.page(page)
    except PageNotAnInteger:
        posts = paginator.page(1)
    except EmptyPage:
        posts = paginator.page(paginator.num_pages)


    context = {
         'posts' : posts
    }
    return render(request,'user/feed.html',context)
    


# Follower Model

@login_required 
def circle(request):
    
    users_list = User.objects.all().exclude(id=(request.user).id)
    page = request.GET.get('page', 1)

    paginator = Paginator(users_list, 10)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)

    context = {
        'users' : users ,
        'users_list' : users_list,
        'owner' : Circle.objects.get(owner=request.user).owner,
        'items' : Circle.objects.get(owner=request.user).friends.all(),
    }

    return render(request,'user/circle.html',context)

@login_required    
def follow(request,id):
    circle = Circle.objects.get(owner=request.user)
    friends = circle.friends.all()
    target = User.objects.get(id=id)

    circle.friends.add(target)
    messages.info(request, f'You started following {target}')
    return redirect("circle")     
         
@login_required    
def unfollow(request,id):
    circle = Circle.objects.get(owner=request.user)
    friends = circle.friends.all()
    target = User.objects.get(id=id)
    context = {
        "target": target
    }
    if request.method == "POST":
            circle.friends.remove(target)
            messages.warning(request, f"You have unfollowed {target}")  
            return redirect("circle")      
                
    return render(request, "user/unfollow.html", context)    





# User Report

@staff_member_required
def report(request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        users = User.objects.all()

        # Defining Column Names
        ws['A1'] = str('Username') 
        ws['B1'] = str('Email') 
        ws['C1'] = str('No. of Posts')

        for user in range(2,len(users)+2):
            for post in Post.objects.filter(author=users[user-2]):
                ws.cell(row=user, column=1).value = str(users[user-2].username)
                ws.cell(row=user, column=2).value = str(users[user-2].email)
                ws.cell(row=user, column=3).value = str(Post.objects.filter(author_id=users[user-2].id).count())

        wb.save("New Report.xlsx")
        wb.close()
        messages.success(request, f"User Report has been successfully generated.")
        return redirect('social-home')











# If you want to get a confirm page
# @login_required    
# def follow(request,id):
#     circle = Circle.objects.get(owner=request.user)
#     friends = circle.friends.all()
#     target = User.objects.get(id=id)
#     context = {
#         "target": target
#     }
#     if request.method == "POST": 
#     circle.friends.add(target)
#     messages.info(
#         request, f'You started following {target}'
#     )
    
#     return redirect("circle")   
            
#     return render(request, "user/follow.html", context) 